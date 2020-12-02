from openpyxl import load_workbook
from PIL import Image

from django.core.management.base import BaseCommand
from django.conf import settings
from django.utils.text import slugify
from django.utils import timezone
from django.core.mail import EmailMessage


from users.models import Department, Employee, Like
from vote.settings import DATA_DIR, MEDIA_ROOT, EMAIL_HOST_USER

from fpdf import FPDF

excel_file = DATA_DIR+'/employee.xlsx'


class Command(BaseCommand):

    timezone_list = (3, 6, 9, 12)
    day = timezone.now().day
    month = timezone.now().month

    def handle(self, *args, **options):
        """
            Обновляет базу данных и проверяет дату для сбора результатов
        """
        self.set_employment()

        if self.month in self.timezone_list and self.day == 1:
            print('Сбор результатов')
            self.get_vote_result()

        print('Start import from excel_file')
        try:
            wb = load_workbook(excel_file)
            sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

            for count in range(2, sheet.max_row+1):
                raw_password = sheet.cell(row=count, column=5).value
                login = sheet.cell(row=count, column=5).value
                first_name = sheet.cell(row=count, column=1).value
                last_name = sheet.cell(row=count, column=2).value
                slug = slugify(
                    f'{first_name} {last_name}',
                    allow_unicode=True
                    )
                position = sheet.cell(row=count, column=3).value
                try:
                    avatar_url = sheet.cell(row=count, column=6).value
                    avatar_name = avatar_url.split('/')[-1]
                    image = Image.open(r'{}'.format(avatar_url))
                    new_image = image.resize((400, 400))
                    new_image.save(f'{MEDIA_ROOT}{avatar_name}')
                except AttributeError:
                    print(f'File {avatar_url} not found')
                except FileNotFoundError:
                    print(f'File {avatar_url} not found')
                finally:
                    print('Import stopped!!!')
                    break

                dep_name = sheet.cell(row=count, column=4).value
                dep_slug = slugify(dep_name, allow_unicode=True)

                department, create = Department.objects.get_or_create(
                    name=dep_name,
                    slug=dep_slug)

                employee, create = Employee.objects.update_or_create(
                    login=self.validate(login),
                    first_name=first_name,
                    last_name=last_name,
                    slug=slug,
                    position=position,
                    avatar=avatar_name,
                    department=department,
                )
                employee.employment = True
                employee.save()
                print(f'Данные {employee} обновлены!!!\n')
                if create or not employee.password:
                    print(f'create {employee} done!!!')
                    print(f'set password to {employee}')
                    try:
                        employee.set_password(
                            self.validate(raw_password)
                            )
                        employee.save()
                    except TypeError:
                        print(
                            f'''Пароль для {first_name} {last_name} не установлен!!!!!
                            \nПароль должен быть представлен в виде строки'''
                            )
            print('Готово!!!\n')
        except FileNotFoundError:
            print('Указаный Вами файл не найден!!!')

    def reset(self):
        """
            Сбрасывает лайки
        """
        print('Сброс лайков')
        Like.objects.all().delete()
        employee_list = Employee.objects.all()
        for employee in employee_list:
            employee.like_count = 0
            employee.save()

    def get_vote_result(self):
        """
            Собирает результат голосований и отправляет на почту
        """
        data = [['  '*10+'отдел', '  '*10+'сотрудник', '  '*10+'голоса']]

        for department in Department.objects.all():
            user_list = department.users.filter(
                is_superuser=False).order_by('-like_count')

            if len(user_list) > 0 and user_list[0].like_count > 0:
                new_list = [department.__str__(),
                            user_list[0].__str__(),
                            str(user_list[0].like_count)]
                data.append(new_list)

        self.create_pdf_report(data)
        self.send_report('load_from_excel.pdf')
        self.reset()

    def create_pdf_report(self, data):
        spacing = 1
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font('DejaVu', '',
                     'static/fonts/DejaVuSansCondensed.ttf', uni=True)
        pdf.set_font('DejaVu', '', 11)
        col_width = pdf.w / 3.3
        row_height = 10
        for row in data:
            for item in row:
                pdf.cell(col_width, row_height*spacing,
                         txt=item, border=1)
            pdf.ln(row_height*spacing)

        pdf.output('load_from_excel.pdf')

    def set_employment(self):
        """
        Ставит значение employment в False
        """
        userlist = Employee.objects.filter(is_superuser=False)
        for user in userlist:
            user.employment = False
            user.save()

    def validate(self, not_validate_login):
        if len(not_validate_login) > 10:
            return not_validate_login[-1:-11:-1][::-1]
        return not_validate_login

    def send_report(self, path_to_filename):
        message = f'Результат голосования {timezone.now().date()}'
        email = EmailMessage(
                             'Результат голосований',
                             message,
                             EMAIL_HOST_USER,
                             [settings.RESULT_TO_EMAIL],
                             )
        email.attach_file(path_to_filename)
        email.send()
